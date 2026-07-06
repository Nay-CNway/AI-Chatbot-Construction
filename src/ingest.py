import openpyxl
import os

file_path = r"C:\Users\naych\AI-Chatbot-Construction\past_projects"

SUMMARY_SHEET_NAMES = [
    "Summary",
    "Summary Of Cost",
    "ALL SUM",
    "SUM",
    "Total Summary"
]

MANUAL_DATA = {
    "project_1_one_story": {
        "grand_total": 307938673.56,
        "floor_area": None,   
        "pae_rate": 183546.94
    },
    "project_2_two_half_storey_kannerlann": {
        "grand_total": 486400000,
        "floor_area": None,   
        "pae_rate": 90969.55
    },
    "project_4_AMT": {
        "grand_total": 268280000,
        "floor_area": None,   
        "pae_rate": None      
    },
    "project_5_NPT": {
        "grand_total": 152808000,
        "floor_area": None,   
        "pae_rate": None     
    },
    "project_6_UYW": {
        "grand_total": 486400000,
        "floor_area": None,   
        "pae_rate": 90969.55
    },
}

all_projects = []

for project_folder in os.listdir(file_path):
    project_path = os.path.join(file_path,project_folder)

    excel_files = [f for f in os.listdir(project_path) if f.endswith(".xlsx")]
    if not excel_files:
        print(f"⚠️ No Excel file found in {project_folder}")
        continue

    excel_path = os.path.join(project_path, excel_files[0])
    
    try:
        workbook = openpyxl.load_workbook(excel_path,data_only= True)
        summary_sheet = None
        found_sheet_name = None
        for name in SUMMARY_SHEET_NAMES:
            if name in workbook.sheetnames:
                summary_sheet = workbook[name]
                found_sheet_name = name
                break

        grand_total = None
        floor_area = None
        pae_rate = None

        if summary_sheet:
            for row in summary_sheet.iter_rows(values_only= True):
                label_col1 = row[0] if len(row) > 0 else None
                label_col2 = row[1] if len(row) > 1 else None
                value_col2 = row[2] if len(row) > 2 else None

                label = label_col1 or label_col2

                if label in ["Grand Total Cost", "Grand Total", "Total Cost", "Say"]:
                    grand_total = value_col2
                elif label in ["Floor Area", "Total Floor Area", "Total Area"]:
                    floor_area = value_col2
                elif label in ["PAE", "PAE Rate", "Rate per sq ft"]:
                    pae_rate = value_col2

        if project_folder in MANUAL_DATA:
            manual = MANUAL_DATA[project_folder]
            grand_total = grand_total or manual["grand_total"]
            floor_area = floor_area or manual["floor_area"]
            pae_rate = pae_rate or manual["pae_rate"]

        all_projects.append({
            "project": project_folder,
            "grand_total": grand_total,
            "floor_area": floor_area,
            "pae_rate" : pae_rate
        })

    except Exception as e:
        print(f"⚠️ Could not read {project_folder}: {e}")
    
print("\n=== ALL PROJECTS SUMMARY ===\n")
for p in all_projects:
    print(f"Project:     {p['project']}")
    print(f"  Grand Total: {p['grand_total']} MMK")
    print(f"  Floor Area:  {p['floor_area']} sq ft")
    print(f"  PAE Rate:    {p['pae_rate']} MMK/sq ft")
    print()


import os

output_folder = r"C:\Users\naych\AI-Chatbot-Construction\data\project_texts"
os.makedirs(output_folder, exist_ok= True)

for p in all_projects:
    project_name = p['project']
    grand_total = p['grand_total']
    floor_area = p['floor_area']
    pae_rate = p['pae_rate']

    text_content = f"""Project: {project_name}
Building Cost Estimation Data for Creative Paradise Construction

Grand Total Cost: {grand_total:,.0f} MMK
Floor Area: {floor_area if floor_area else 'Not Available'} sq ft
Cost per sq ft (PAE Rate): {pae_rate if pae_rate else 'Not Available'} MMK/sq ft

This is a real completed project by Creative Paradise Construction and Decoration Co., Ltd based in Myanmar.
The cost includes material, labour, machinery, transportation, preliminary and supervision charges.
"""
    
    output_path = os.path.join(output_folder, f"{project_name}.txt")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(text_content)
    
    print(f"✅ Saved: {project_name}.txt")

print(f"\n✅ All project text files saved to: {output_folder}")